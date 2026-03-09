import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Trang tính1"

# === Data ===
ten_van_ban = "Thông tư quy định về hệ thống kiểm soát nội bộ của ngân hàng thương mại, chi nhánh ngân hàng nước ngoài"
so_van_ban = "83/2025/TT-NHNN"
ngay_hieu_luc = date(2026, 7, 1)

obligations = [
    {
        "dieu": "4",
        "nghia_vu": (
            "Điều 4. Yêu cầu đối với hệ thống kiểm soát nội bộ\n"
            "Hệ thống kiểm soát nội bộ phải:\n"
            "a) Tuân thủ Luật Các tổ chức tín dụng, Thông tư này và các quy định pháp luật có liên quan; xử lý, khắc phục theo yêu cầu của NHNN, kiểm toán độc lập;\n"
            "b) Phù hợp với quy mô, tính chất và mức độ phức tạp trong hoạt động kinh doanh;\n"
            "c) Có đủ nguồn lực về tài chính, con người, hệ thống thông tin;\n"
            "d) Xây dựng, duy trì văn hóa kiểm soát;\n"
            "đ) Có hệ thống thông tin quản lý đáp ứng quy định tại Điều 8;\n"
            "e) Có 03 tuyến bảo vệ độc lập. Ý kiến thảo luận, kết luận về KSNB trong cuộc họp phải được ghi lại bằng văn bản."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Rà soát và đảm bảo hệ thống KSNB tuân thủ đầy đủ Thông tư 83/2025;\n"
            "- Xây dựng và duy trì 03 tuyến bảo vệ độc lập (tuyến 1: bộ phận tạo ra rủi ro; tuyến 2: tuân thủ + quản lý rủi ro; tuyến 3: kiểm toán nội bộ);\n"
            "- Ghi nhận bằng văn bản ý kiến, kết luận liên quan đến KSNB trong cuộc họp HĐQT, BKS;\n"
            "- Đảm bảo đủ nguồn lực tài chính, nhân lực, hạ tầng CNTT cho KSNB."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "5",
        "nghia_vu": (
            "Điều 5. Cơ chế, chính sách, quy trình, quy định nội bộ\n"
            "Quy định nội bộ phải:\n"
            "a) Phù hợp với Thông tư này và pháp luật;\n"
            "b) Phân cấp thẩm quyền rõ ràng theo tiêu chí quy mô, hạn mức rủi ro;\n"
            "c) Phân tách chức năng, nhiệm vụ để ngăn ngừa xung đột lợi ích; không để một cá nhân chi phối toàn bộ giao dịch;\n"
            "d) Phân cấp trách nhiệm quản lý tài sản rõ ràng;\n"
            "đ) Ban hành chuẩn mực đạo đức nghề nghiệp (do HĐQT/TGĐ ban hành);\n"
            "e) Được đánh giá định kỳ và sửa đổi, bổ sung khi cần;\n"
            "g) Có quy định về kiểm soát trụ sở chính đối với chi nhánh, đơn vị phụ thuộc."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Rà soát toàn bộ quy định nội bộ hiện hành để đảm bảo phù hợp Thông tư 83/2025;\n"
            "- Đảm bảo phân tách chức năng thẩm định tín dụng với phê duyệt, quan hệ khách hàng;\n"
            "- HĐQT ban hành/cập nhật chuẩn mực đạo đức nghề nghiệp cho toàn bộ CBNV;\n"
            "- TGĐ ban hành quy trình nội bộ vận hành; HĐQT ban hành quy định về tổ chức, quản trị;\n"
            "- Thiết lập cơ chế kiểm soát trụ sở chính đối với các chi nhánh, đơn vị phụ thuộc;\n"
            "- Đánh giá định kỳ và cập nhật quy định nội bộ khi pháp luật thay đổi."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "6",
        "nghia_vu": (
            "Điều 6. Cơ cấu tổ chức thực hiện hệ thống kiểm soát nội bộ\n"
            "NHTM phải:\n"
            "a) Thành lập các ủy ban theo Điều 50 Luật TCTD (Ủy ban quản lý rủi ro, Ủy ban nhân sự); mỗi ủy ban có trên 1/2 thành viên không phải người điều hành;\n"
            "b) Thành lập Hội đồng rủi ro (họp ít nhất mỗi quý 1 lần), Hội đồng ALCO (họp ít nhất mỗi quý 1 lần), Hội đồng quản lý vốn (họp ít nhất 6 tháng 1 lần);\n"
            "c) Có Bộ phận tuân thủ và Bộ phận quản lý rủi ro (tuyến bảo vệ thứ hai);\n"
            "d) TGĐ quyết định quy chế làm việc của các Hội đồng (chức năng, nhiệm vụ, cơ chế ra quyết định, cơ chế họp)."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Rà soát cơ cấu tổ chức các ủy ban của HĐQT, đảm bảo đủ thành viên không điều hành;\n"
            "- Thiết lập/rà soát quy chế làm việc của Hội đồng rủi ro, Hội đồng ALCO, Hội đồng quản lý vốn;\n"
            "- Đảm bảo tần suất họp định kỳ: Hội đồng rủi ro và ALCO tối thiểu hàng quý; Hội đồng quản lý vốn tối thiểu 6 tháng;\n"
            "- Bộ phận tuân thủ và bộ phận QLRR hoạt động độc lập, không xung đột lợi ích."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "8",
        "nghia_vu": (
            "Điều 8. Hệ thống thông tin quản lý\n"
            "Ngân hàng phải thiết lập hệ thống thông tin quản lý (MIS) bao gồm tối thiểu:\n"
            "a) Cơ cấu tổ chức quản lý, vận hành MIS và quy định nội bộ về trách nhiệm;\n"
            "b) Quy trình thu thập, xử lý, lưu trữ, cung cấp thông tin;\n"
            "c) Dữ liệu phục vụ quản lý, điều hành (bao gồm dữ liệu rủi ro theo Điều 24);\n"
            "d) Các báo cáo nội bộ (hoạt động kiểm soát, quản lý rủi ro, ICAAP, kiểm toán nội bộ);\n"
            "đ) Hạ tầng CNTT phù hợp; hệ thống dự phòng an toàn.\n"
            "MIS phải được rà soát, đánh giá tối thiểu hằng năm và đột xuất."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành/rà soát quy định nội bộ về trách nhiệm cá nhân, bộ phận trong quản lý, vận hành MIS;\n"
            "- Đảm bảo MIS cung cấp đầy đủ, kịp thời thông tin cho HĐQT, BKS, TGĐ và các bộ phận;\n"
            "- Có cơ chế báo cáo vi phạm pháp luật, vi phạm quy định nội bộ một cách bảo mật;\n"
            "- Hệ thống dự phòng đảm bảo dữ liệu an toàn, không gián đoạn;\n"
            "- Rà soát, đánh giá và nâng cấp MIS tối thiểu hàng năm."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "9",
        "nghia_vu": (
            "Điều 9. Báo cáo Ngân hàng Nhà nước về hệ thống kiểm soát nội bộ\n"
            "Ngân hàng phải lập và gửi NHNN các báo cáo hằng năm:\n"
            "a) Báo cáo kết quả tự kiểm tra, đánh giá hoạt động kiểm soát (Phụ lục I) - trong 90 ngày từ ngày kết thúc năm tài chính; HĐQT phê duyệt;\n"
            "b) Báo cáo quản lý rủi ro (Phụ lục II) - trong 90 ngày; TGĐ phê duyệt;\n"
            "c) Báo cáo ICAAP (Phụ lục III) - trong 90 ngày; TGĐ phê duyệt;\n"
            "d) Báo cáo kiểm toán nội bộ (Phụ lục IV) - trong 60 ngày; BKS phê duyệt; báo cáo đột xuất trong 07 ngày làm việc sau khi kết thúc kiểm toán đột xuất."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Lập lịch nội bộ để đảm bảo các báo cáo KSNB gửi NHNN đúng hạn;\n"
            "- Báo cáo hoạt động kiểm soát, QLRR, ICAAP: gửi trong 90 ngày kể từ 31/12 hàng năm;\n"
            "- Báo cáo kiểm toán nội bộ định kỳ: gửi trong 60 ngày kể từ 31/12;\n"
            "- Báo cáo kiểm toán nội bộ đột xuất: gửi trong 07 ngày làm việc sau khi kết thúc;\n"
            "- Phân công thẩm quyền phê duyệt đúng: HĐQT (báo cáo kiểm soát), TGĐ (QLRR, ICAAP), BKS (kiểm toán nội bộ)."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "10",
        "nghia_vu": (
            "Điều 10. Lưu trữ hồ sơ, tài liệu về hệ thống kiểm soát nội bộ\n"
            "Ngân hàng phải:\n"
            "a) Có quy định nội bộ về quản lý, lưu trữ hồ sơ, tài liệu về hệ thống KSNB;\n"
            "b) Tuân thủ quy định của NHNN về thời hạn bảo quản hồ sơ, tài liệu;\n"
            "c) Lưu trữ đầy đủ để cung cấp theo yêu cầu của kiểm toán nội bộ, kiểm toán độc lập, cơ quan có thẩm quyền."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành/rà soát quy định nội bộ về quản lý, lưu trữ hồ sơ KSNB;\n"
            "- Đảm bảo thời hạn lưu trữ tuân thủ quy định của NHNN;\n"
            "- Đảm bảo hồ sơ sẵn sàng cung cấp khi có yêu cầu từ kiểm toán nội bộ, kiểm toán độc lập, NHNN và cơ quan có thẩm quyền."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "11",
        "nghia_vu": (
            "Điều 11. Yêu cầu của hoạt động kiểm soát\n"
            "Hoạt động kiểm soát phải:\n"
            "1. Thực hiện đối với tất cả hoạt động, quy trình, cá nhân, bộ phận;\n"
            "2. Hạch toán kế toán tuân thủ chuẩn mực và chế độ kế toán; kiểm tra, đối chiếu thường xuyên;\n"
            "3. Có biện pháp phòng ngừa, xử lý kịp thời sai phạm;\n"
            "4. Phân bổ nhân lực phù hợp, có phương án nhân sự thay thế;\n"
            "5. Trụ sở chính giám sát, kiểm soát chi nhánh, đơn vị phụ thuộc;\n"
            "6. Báo cáo nội bộ về hoạt động kiểm soát hằng năm và đột xuất;\n"
            "7. Có quy định về kiểm soát hoạt động điện tử."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Đảm bảo hoạt động kiểm soát bao phủ tất cả hoạt động, quy trình;\n"
            "- Có cơ chế kiểm tra, đối chiếu hạch toán kế toán định kỳ;\n"
            "- Xây dựng phương án nhân sự thay thế cho các vị trí kiểm soát;\n"
            "- Trụ sở chính quyết định chức năng, nhiệm vụ, lương thưởng, kỷ luật nhân sự kiểm soát tại chi nhánh;\n"
            "- Lập báo cáo nội bộ về hoạt động kiểm soát hằng năm;\n"
            "- Ban hành quy định kiểm soát các hoạt động thực hiện bằng phương tiện điện tử."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "14",
        "nghia_vu": (
            "Điều 14. Trách nhiệm trong hoạt động kiểm soát\n"
            "Bộ phận tuân thủ (tối thiểu) phải:\n"
            "a) Đầu mối xác định các vấn đề tuân thủ;\n"
            "b) Đánh giá định kỳ tính thích hợp, tuân thủ pháp luật của quy định nội bộ;\n"
            "c) Báo cáo TGĐ về tình hình tuân thủ, vi phạm nghiêm trọng, thay đổi quy định;\n"
            "d) Trực tiếp báo cáo HĐQT trong trường hợp cần thiết (kể cả khi TGĐ vi phạm);\n"
            "đ) Theo dõi, kiểm tra các cá nhân, bộ phận về tuân thủ pháp luật, quy định nội bộ;\n"
            "e) Hỗ trợ xây dựng, rà soát quy định nội bộ; thông báo thay đổi pháp luật."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Đảm bảo Bộ phận tuân thủ thực hiện đầy đủ 07 nhiệm vụ tối thiểu theo Điều 14;\n"
            "- Bộ phận tuân thủ có cơ chế báo cáo thẳng lên HĐQT khi phát hiện vi phạm nghiêm trọng của TGĐ;\n"
            "- Xây dựng hệ thống theo dõi thay đổi pháp luật và thông báo kịp thời đến các bộ phận liên quan;\n"
            "- TGĐ tổ chức thực hiện hoạt động kiểm soát, duy trì văn hóa kiểm soát và chuẩn mực đạo đức nghề nghiệp;\n"
            "- TGĐ xử lý vi phạm quy định nội bộ và chuyển cơ quan thẩm quyền xử lý vi phạm pháp luật."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "15",
        "nghia_vu": (
            "Điều 15. Hoạt động kiểm soát trong hoạt động cấp tín dụng\n"
            "Hoạt động cấp tín dụng phải kiểm soát xung đột lợi ích:\n"
            "- Bộ phận thẩm định tín dụng phải độc lập với bộ phận: (a) phê duyệt cấp tín dụng; (b) kiểm soát hạn mức rủi ro, quản lý nợ xấu; (c) quan hệ khách hàng.\n"
            "- Cấp tín dụng qua phương tiện điện tử phải phân định trách nhiệm rõ ràng giữa bộ phận xây dựng hệ thống và bộ phận thẩm định, quyết định; khi phát sinh rủi ro phải xác định được từng cá nhân, bộ phận chịu trách nhiệm."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Đảm bảo phân tách chức năng: thẩm định tín dụng độc lập với phê duyệt, quan hệ khách hàng, kiểm soát hạn mức;\n"
            "- Rà soát cơ cấu tổ chức quy trình cấp tín dụng để loại bỏ xung đột lợi ích;\n"
            "- Đối với cấp tín dụng điện tử: phân định rõ trách nhiệm giữa IT/vận hành hệ thống với thẩm định tín dụng;\n"
            "- Có cơ chế xác định, quy trách nhiệm và xử lý nhanh khi xảy ra rủi ro trong cấp tín dụng điện tử."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "16",
        "nghia_vu": (
            "Điều 16. Hoạt động kiểm soát trong giao dịch tự doanh\n"
            "Giao dịch tự doanh phải đảm bảo:\n"
            "a) Bộ phận chuyên trách giao dịch tự doanh độc lập với bộ phận kiểm soát và thanh toán giao dịch;\n"
            "b) Giao dịch thực hiện trong hạn mức; đàm thoại điện thoại phải ghi âm và lưu tối thiểu 02 tháng; hệ thống máy tính tự động ghi nhận ngày giờ, mã giao dịch;\n"
            "c) Giá cả giao dịch phải được kiểm tra độc lập, phù hợp giá thị trường;\n"
            "d) Thanh toán giao dịch: gửi, nhận xác nhận giao dịch; nội dung xác nhận đầy đủ; xử lý kịp thời chênh lệch."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Đảm bảo phân tách 3 chức năng: front office (giao dịch), middle office (kiểm soát), back office (thanh toán);\n"
            "- Lưu trữ ghi âm đàm thoại giao dịch tự doanh tối thiểu 02 tháng;\n"
            "- Hệ thống CNTT tự động ghi nhận ngày, giờ, mã số giao dịch; không cho phép giao dịch viên thay đổi;\n"
            "- Có cơ chế kiểm tra giá độc lập so với giá thị trường;\n"
            "- Quy trình xác nhận giao dịch và xử lý chênh lệch phải được quy định rõ ràng."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "18",
        "nghia_vu": (
            "Điều 18. Quy định nội bộ về hoạt động quản lý rủi ro\n"
            "Ngân hàng phải có quy định nội bộ về QLRR, tối thiểu bao gồm:\n"
            "a) Xây dựng, ban hành và thực hiện chính sách QLRR;\n"
            "b) Hạn mức rủi ro đối với từng loại rủi ro trọng yếu (phương pháp xây dựng, phân bổ, xử lý vi phạm);\n"
            "c) Nhận dạng, đo lường, theo dõi và kiểm soát các rủi ro trọng yếu;\n"
            "d) Cơ chế báo cáo nội bộ về QLRR;\n"
            "đ) Quản lý rủi ro sản phẩm mới, thị trường mới;\n"
            "e) Kiểm tra sức chịu đựng;\n"
            "g) Quản lý dữ liệu rủi ro;\n"
            "h) Quy trình ICAAP."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành/cập nhật quy định nội bộ về QLRR đảm bảo đầy đủ 09 nội dung tối thiểu;\n"
            "- Quy định nội bộ phải phù hợp chiến lược kinh doanh, văn hóa kiểm soát, nguồn nhân lực, hạ tầng CNTT;\n"
            "- Đảm bảo cơ chế báo cáo kịp thời vi phạm, trạng thái rủi ro cho HĐQT, BKS;\n"
            "- Có cơ chế xử lý vi phạm về QLRR."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "19",
        "nghia_vu": (
            "Điều 19. Chính sách quản lý rủi ro\n"
            "Chính sách QLRR do HĐQT ban hành, sửa đổi. Tối thiểu bao gồm:\n"
            "a) Khẩu vị rủi ro: tỷ lệ an toàn vốn mục tiêu, chỉ tiêu thu nhập (ROE, RAROC), chỉ tiêu định tính khác;\n"
            "b) Danh sách các rủi ro trọng yếu;\n"
            "c) Chiến lược QLRR từng loại rủi ro trọng yếu.\n"
            "Chính sách phải: lập cho 03-05 năm tiếp theo; đánh giá định kỳ tối thiểu hằng năm; có tính kế thừa, liên tục."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- HĐQT ban hành và định kỳ đánh giá lại (ít nhất hàng năm) chính sách QLRR;\n"
            "- Chính sách QLRR xác định rõ khẩu vị rủi ro (tỷ lệ an toàn vốn mục tiêu, ROE mục tiêu, RAROC);\n"
            "- Xác định danh sách rủi ro trọng yếu và chiến lược quản lý từng loại;\n"
            "- Chính sách phải cho giai đoạn 03-05 năm, có tính kế thừa qua các chu kỳ kinh tế;\n"
            "- Chính sách phù hợp với năng lực tài chính, vốn tự có và nguồn tăng vốn."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "20",
        "nghia_vu": (
            "Điều 20. Hạn mức rủi ro\n"
            "Hạn mức rủi ro do TGĐ ban hành. Hạn mức phải:\n"
            "a) Tuân thủ các hạn chế an toàn theo Luật TCTD và quy định NHNN;\n"
            "b) Có hạn mức kiểm soát các rủi ro trọng yếu;\n"
            "c) Phù hợp với khẩu vị rủi ro, chiến lược QLRR;\n"
            "d) Rà soát, đánh giá lại ít nhất hằng năm hoặc khi có thay đổi lớn; nếu nới lỏng hạn mức, TGĐ phải báo cáo HĐQT sau khi điều chỉnh;\n"
            "đ) Được phổ biến đến các cá nhân, bộ phận liên quan."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- TGĐ ban hành hệ thống hạn mức rủi ro cho từng loại rủi ro trọng yếu;\n"
            "- Rà soát hạn mức rủi ro ít nhất hàng năm hoặc khi có thay đổi lớn về trạng thái rủi ro;\n"
            "- Khi nới lỏng hạn mức: TGĐ phải báo cáo HĐQT ngay sau khi điều chỉnh;\n"
            "- Phổ biến hạn mức rủi ro đến tất cả cá nhân, bộ phận có liên quan;\n"
            "- Áp dụng hạn mức thận trọng hơn khi một hoạt động có nhiều loại rủi ro khác nhau."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "21",
        "nghia_vu": (
            "Điều 21. Quản lý rủi ro đối với sản phẩm mới, hoạt động trong thị trường mới\n"
            "Ngân hàng phải:\n"
            "a) Có quy định nội bộ xác định tiêu chí sản phẩm mới, thị trường mới;\n"
            "b) HĐQT phê duyệt chủ trương; TGĐ phê duyệt kế hoạch cung cấp sản phẩm mới;\n"
            "c) Kế hoạch phải được Bộ phận QLRR thẩm định về rủi ro, biện pháp kiểm soát;\n"
            "d) Xác định quy mô, thời gian thử nghiệm; đánh giá kết quả trước khi chính thức triển khai;\n"
            "e) Khi chính thức triển khai: ban hành quy định, quy trình và thực hiện QLRR đối với sản phẩm/thị trường mới."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành tiêu chí xác định sản phẩm mới, thị trường mới trong quy định nội bộ;\n"
            "- Quy trình phê duyệt 2 cấp: HĐQT phê duyệt chủ trương, TGĐ phê duyệt kế hoạch chi tiết;\n"
            "- Bộ phận QLRR phải thẩm định rủi ro trước khi TGĐ phê duyệt kế hoạch;\n"
            "- Thực hiện thử nghiệm, đánh giá kết quả trước khi triển khai chính thức;\n"
            "- Ban hành quy định, quy trình cho sản phẩm mới khi chính thức triển khai."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "23",
        "nghia_vu": (
            "Điều 23. Kiểm tra sức chịu đựng (Stress Testing)\n"
            "Ngân hàng phải thực hiện kiểm tra sức chịu đựng về vốn và cho các rủi ro trọng yếu:\n"
            "a) Rủi ro tín dụng và thị trường: tối thiểu hằng năm (từ 01/01/2028);\n"
            "b) Rủi ro lãi suất trên sổ ngân hàng: tối thiểu hằng quý;\n"
            "c) Rủi ro thanh khoản: tối thiểu hằng quý;\n"
            "d) Vốn: tối thiểu hằng năm.\n"
            "Phải ban hành quy định nội bộ về stress testing; áp dụng ít nhất 1 trong 3 phương pháp (phân tích độ nhạy, phân tích kịch bản, reverse stress testing). Kết quả stress testing phải được sử dụng trong xây dựng chiến lược kinh doanh, chính sách QLRR."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành quy định nội bộ về stress testing (loại, tần suất, phương pháp luận, phạm vi, quy trình);\n"
            "- Thực hiện stress testing thanh khoản và IRRBB hàng quý;\n"
            "- Thực hiện stress testing tín dụng, thị trường và vốn hàng năm (tín dụng/thị trường từ 01/01/2028);\n"
            "- Áp dụng tối thiểu 01 trong 03 phương pháp theo quy định;\n"
            "- Sử dụng kết quả stress testing để điều chỉnh chiến lược kinh doanh, chính sách QLRR, kế hoạch dự phòng."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "24",
        "nghia_vu": (
            "Điều 24. Quản lý dữ liệu rủi ro (từ 01/01/2028)\n"
            "Ngân hàng quản lý dữ liệu rủi ro theo quy định nội bộ, tối thiểu:\n"
            "1. Dữ liệu rủi ro phải chính xác (accuracy) và toàn vẹn (integrity); nhất quán khái niệm; có kênh báo cáo nội bộ và kế hoạch khắc phục dữ liệu kém chất lượng;\n"
            "2. Thu thập và tổng hợp được dữ liệu rủi ro từ tất cả rủi ro trọng yếu;\n"
            "3. Tổng hợp dữ liệu nhanh chóng, kịp thời;\n"
            "4. Đáp ứng yêu cầu các loại báo cáo QLRR khác nhau."
        ),
        "hanh_dong": (
            "TCB phải (trước 01/01/2028):\n"
            "- Ban hành quy định nội bộ về quản lý dữ liệu rủi ro;\n"
            "- Xây dựng hạ tầng dữ liệu đảm bảo chính xác, toàn vẹn, nhất quán khái niệm;\n"
            "- Thiết lập quy trình thu thập, tổng hợp dữ liệu từ tất cả rủi ro trọng yếu;\n"
            "- Có kế hoạch khắc phục khi phát hiện dữ liệu kém chất lượng;\n"
            "- Đảm bảo hệ thống có khả năng tổng hợp dữ liệu nhanh chóng cho các loại báo cáo QLRR."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "26",
        "nghia_vu": (
            "Điều 26. Chiến lược và hạn mức rủi ro tín dụng\n"
            "Ngân hàng phải ban hành chiến lược QLRR tín dụng, tối thiểu:\n"
            "a) Mục tiêu chất lượng tín dụng (tỷ lệ nợ xấu mục tiêu, tỷ lệ cấp tín dụng xấu mục tiêu) theo sản phẩm, khách hàng, ngành kinh tế;\n"
            "b) Nguyên tắc tính chi phí bù đắp rủi ro trong lãi suất, định giá sản phẩm tín dụng;\n"
            "c) Nguyên tắc áp dụng biện pháp giảm thiểu rủi ro tín dụng;\n"
            "d) Nguyên tắc kiểm tra sức chịu đựng về rủi ro tín dụng.\n"
            "Hạn mức rủi ro tín dụng tối thiểu bao gồm hạn mức theo sản phẩm, khách hàng, ngành kinh tế, hình thức bảo đảm."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành chiến lược QLRR tín dụng với các mục tiêu chất lượng tín dụng cụ thể theo sản phẩm/khách hàng/ngành;\n"
            "- Tích hợp chi phí bù đắp rủi ro tín dụng vào cơ chế định giá sản phẩm tín dụng;\n"
            "- Thiết lập hạn mức rủi ro tín dụng theo sản phẩm, khách hàng, ngành kinh tế, hình thức bảo đảm;\n"
            "- Quy định rõ thẩm quyền phê duyệt biện pháp giảm thiểu rủi ro tín dụng."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "27",
        "nghia_vu": (
            "Điều 27. Hệ thống xếp hạng tín dụng nội bộ\n"
            "Ngân hàng phải có hệ thống XHTN nội bộ theo quy định NHNN. Tối thiểu:\n"
            "a) Tiêu chí đánh giá khả năng trả nợ (bao gồm các yếu tố kinh tế - xã hội vĩ mô);\n"
            "b) Hệ thống thông tin để vận hành;\n"
            "c) Đánh giá độc lập tối thiểu hằng năm để rà soát, sửa đổi;\n"
            "d) Đầy đủ thông tin cung cấp theo yêu cầu của kiểm toán, NHNN khi kiểm tra, thanh tra."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Duy trì hệ thống XHTN nội bộ tuân thủ quy định NHNN về phân loại tài sản;\n"
            "- Thực hiện đánh giá độc lập hệ thống XHTN tối thiểu hằng năm (bởi bộ phận không xây dựng/sử dụng hệ thống);\n"
            "- Lưu trữ đầy đủ hồ sơ, tài liệu về hệ thống XHTN để sẵn sàng cung cấp khi thanh tra, kiểm tra."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "28",
        "nghia_vu": (
            "Điều 28. Đo lường, theo dõi và kiểm soát rủi ro tín dụng\n"
            "Ngân hàng phải:\n"
            "a) Phân loại nợ, trích lập và sử dụng dự phòng rủi ro theo quy định;\n"
            "b) Theo dõi, đánh giá rủi ro từng khoản và toàn bộ danh mục cấp tín dụng;\n"
            "c) Kiểm soát trạng thái rủi ro tín dụng theo hạn mức được phân bổ;\n"
            "d) Giám sát từ xa và kiểm tra tại chỗ với khách hàng; tần suất theo quy định nội bộ;\n"
            "đ) Xây dựng tiêu chí đánh giá suy giảm chất lượng tín dụng; có cơ chế cảnh báo sớm."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Tuân thủ quy định về phân loại nợ và trích lập dự phòng rủi ro;\n"
            "- Xây dựng tiêu chí, phương pháp đánh giá suy giảm chất lượng tín dụng;\n"
            "- Có hệ thống cảnh báo sớm khi chất lượng tín dụng có nguy cơ suy giảm;\n"
            "- Quy định tần suất giám sát từ xa và kiểm tra tại chỗ đối với từng loại khách hàng;\n"
            "- Kiểm soát chặt dư nợ theo hạn mức rủi ro tín dụng đã phân bổ."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "31",
        "nghia_vu": (
            "Điều 31. Quản lý cấp tín dụng\n"
            "Ngân hàng phải:\n"
            "a) Quy định trách nhiệm, thẩm quyền lập, lưu trữ hồ sơ tín dụng đầy đủ theo pháp luật;\n"
            "b) Giải ngân phù hợp với mục đích vay, loại hình cấp tín dụng;\n"
            "c) Kiểm tra, giám sát sau giải ngân: kiểm tra sử dụng vốn, đánh giá khả năng trả nợ, quản lý TSĐB, theo dõi lịch trả nợ;\n"
            "d) Lưu trữ hồ sơ tín dụng, thông tin khả năng trả nợ, lịch sử trả nợ theo quy định pháp luật."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Phân cấp rõ ràng thẩm quyền lập, phê duyệt, lưu trữ hồ sơ tín dụng;\n"
            "- Kiểm tra sử dụng vốn sau giải ngân theo hợp đồng và quy định pháp luật;\n"
            "- Theo dõi lịch trả nợ, nhắc nhở khách hàng và báo cáo kịp thời khi có nguy cơ không trả nợ;\n"
            "- Quản lý TSĐB theo quy định tại Điều 33;\n"
            "- Lưu trữ đầy đủ hồ sơ tín dụng theo quy định pháp luật."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "33",
        "nghia_vu": (
            "Điều 33. Quản lý tài sản bảo đảm\n"
            "Ngân hàng phải có quy định nội bộ về quản lý TSĐB, tối thiểu:\n"
            "a) Xác định các loại TSĐB được chấp nhận;\n"
            "b) Phương pháp xác định giá trị tài sản (theo quy định định giá hoặc thuê tổ chức thẩm định);\n"
            "c) Đánh giá mức độ biến động giá trị TSĐB định kỳ hoặc đột xuất (tài sản biến động nhiều phải đánh giá thường xuyên hơn);\n"
            "d) Quy định về tiếp nhận, bảo quản an toàn TSĐB."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành/cập nhật quy định nội bộ về quản lý TSĐB đầy đủ 04 nội dung tối thiểu;\n"
            "- Đánh giá lại giá trị TSĐB định kỳ và đột xuất; tài sản biến động giá nhiều phải đánh giá thường xuyên hơn;\n"
            "- Có quy trình tiếp nhận, lưu giữ, bảo quản TSĐB an toàn."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "34",
        "nghia_vu": (
            "Điều 34. Báo cáo nội bộ về rủi ro tín dụng\n"
            "Tối thiểu hằng quý hoặc đột xuất, ngân hàng có báo cáo nội bộ về rủi ro tín dụng, bao gồm:\n"
            "- Chất lượng tín dụng theo đối tượng, ngành kinh tế;\n"
            "- Khoản cấp tín dụng có vấn đề và biện pháp xử lý;\n"
            "- Khách hàng/ngành vượt hạn mức và biện pháp kiểm soát;\n"
            "- Giá trị, cơ cấu TSĐB;\n"
            "- Tình hình trích lập, sử dụng dự phòng;\n"
            "- Kết quả stress testing rủi ro tín dụng;\n"
            "- Cảnh báo sớm vi phạm hạn mức;\n"
            "- Đề xuất, kiến nghị."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Lập báo cáo nội bộ về rủi ro tín dụng tối thiểu hàng quý;\n"
            "- Báo cáo phải bao gồm đầy đủ các nội dung theo Điều 34;\n"
            "- Đảm bảo báo cáo được gửi đúng cấp nhận theo quy định nội bộ."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "38",
        "nghia_vu": (
            "Điều 38. Chiến lược và hạn mức rủi ro hoạt động\n"
            "Chiến lược QLRR hoạt động tối thiểu bao gồm:\n"
            "a) Nguyên tắc thực hiện QLRR hoạt động;\n"
            "b) Nguyên tắc sử dụng thuê ngoài, mua bảo hiểm, ứng dụng công nghệ;\n"
            "c) Các trường hợp phải có kế hoạch duy trì hoạt động liên tục (mất tài liệu/CSDL quan trọng, sự cố CNTT, bất khả kháng);\n"
            "d) Nguyên tắc đánh giá rủi ro hoạt động theo kịch bản giả định.\n"
            "Hạn mức rủi ro hoạt động tối thiểu: hạn mức tổn thất tài chính và tổn thất phi tài chính theo 06 nhóm hoạt động."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành chiến lược QLRR hoạt động đầy đủ 04 nội dung tối thiểu;\n"
            "- Phân loại hoạt động thành 06 nhóm theo Khoản 3 Điều 38;\n"
            "- Thiết lập hạn mức tổn thất tài chính và phi tài chính cho từng nhóm hoạt động;\n"
            "- Xác định rõ các trường hợp phải có kế hoạch duy trì hoạt động liên tục."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "39",
        "nghia_vu": (
            "Điều 39. Nhận dạng, đo lường rủi ro hoạt động\n"
            "Ngân hàng phải nhận dạng đầy đủ rủi ro hoạt động trong tất cả giao dịch, sản phẩm, hoạt động, quy trình, hệ thống, rủi ro từ bên thứ ba. Phải áp dụng tối thiểu 02 trong 09 công cụ đo lường rủi ro hoạt động (bao gồm: quản lý sự kiện, thu thập dữ liệu tổn thất, RCSA, khung giám sát kiểm soát, chỉ số KRI/KPI, phân tích kịch bản, đối chiếu/so sánh, BPM, phát hiện kiểm toán).\n"
            "Khi tổn thất vượt hạn mức, phải có biện pháp tăng cường kiểm soát."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Nhận dạng đầy đủ rủi ro hoạt động trong tất cả các giao dịch, sản phẩm, quy trình, hệ thống;\n"
            "- Áp dụng tối thiểu 02 trong 09 công cụ đo lường rủi ro hoạt động theo quy định;\n"
            "- Có biện pháp kiểm soát, tăng cường khi tổn thất thực tế vượt hạn mức rủi ro hoạt động."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "40",
        "nghia_vu": (
            "Điều 40. Quản lý rủi ro hoạt động đối với hoạt động thuê ngoài\n"
            "Ngân hàng phải quản lý hoạt động thuê ngoài tối thiểu:\n"
            "a) Xác định phạm vi và mức độ phụ thuộc vào bên thứ ba;\n"
            "b) Phân cấp thẩm quyền phê duyệt thuê ngoài;\n"
            "c) Thẩm định năng lực bên thứ ba trước khi ký hợp đồng; đánh giá định kỳ;\n"
            "d) Hợp đồng thuê ngoài đầy đủ (phạm vi, quyền nghĩa vụ, xử lý tranh chấp);\n"
            "đ) Lập kế hoạch duy trì hoạt động liên tục cho hoạt động thuê ngoài;\n"
            "e) Thiết lập cơ chế giám sát bên thứ ba trong thời gian thực hiện."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Xây dựng danh mục hoạt động thuê ngoài và mức độ phụ thuộc vào từng bên thứ ba;\n"
            "- Thẩm định năng lực bên thứ ba trước khi ký hợp đồng và đánh giá định kỳ;\n"
            "- Hợp đồng thuê ngoài phải đầy đủ: phạm vi, quyền/nghĩa vụ, bảo mật CSDL, quyền chấm dứt hợp đồng;\n"
            "- Giám sát thường xuyên hoạt động của bên thứ ba;\n"
            "- Có kế hoạch duy trì hoạt động liên tục khi hoạt động thuê ngoài xảy ra sự cố;\n"
            "- Trách nhiệm của TCB với bên liên quan không thay đổi khi thuê ngoài."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "43",
        "nghia_vu": (
            "Điều 43. Kế hoạch duy trì hoạt động liên tục (BCP)\n"
            "Ngân hàng phải xây dựng, phê duyệt, duy trì và cập nhật BCP, tối thiểu:\n"
            "a) Xác định thẩm quyền kích hoạt BCP và vai trò, trách nhiệm các bộ phận;\n"
            "b) Phù hợp với tính chất, quy mô hoạt động;\n"
            "c) Xác định hoạt động, sản phẩm, quy trình trọng yếu cần ưu tiên duy trì;\n"
            "d) Có hệ thống dự phòng nhân sự, địa điểm thay thế, hạ tầng thiết yếu;\n"
            "đ) Đảm bảo phối hợp, trao đổi thông tin kịp thời trong và ngoài ngân hàng;\n"
            "e) Rà soát và thử nghiệm/diễn tập tối thiểu hằng năm."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Xây dựng và ban hành BCP theo đúng các yêu cầu tối thiểu tại Điều 43;\n"
            "- Xác định rõ thẩm quyền kích hoạt BCP và danh sách bộ phận tham gia ứng phó;\n"
            "- Xác định danh sách hoạt động/quy trình trọng yếu cần ưu tiên duy trì;\n"
            "- Có phương án dự phòng nhân sự, địa điểm, hệ thống CNTT;\n"
            "- Thực hiện diễn tập/thử nghiệm BCP tối thiểu hàng năm và điều chỉnh sau diễn tập."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "45",
        "nghia_vu": (
            "Điều 45. Chiến lược và hạn mức rủi ro thanh khoản\n"
            "Chiến lược QLRR thanh khoản tối thiểu bao gồm:\n"
            "a) Nguyên tắc quản lý thanh khoản; duy trì đủ tài sản có tính thanh khoản cao;\n"
            "b) Chiến lược đa dạng hóa nguồn vốn huy động, thời hạn vốn;\n"
            "c) Nguyên tắc kiểm tra sức chịu đựng về thanh khoản.\n"
            "Hạn mức rủi ro thanh khoản tối thiểu: tỷ lệ khả năng chi trả; tỷ lệ dư nợ/tiền gửi; tỷ lệ vốn ngắn hạn cho vay trung dài hạn."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành chiến lược QLRR thanh khoản đầy đủ 03 nội dung tối thiểu;\n"
            "- Thiết lập hạn mức rủi ro thanh khoản đáp ứng quy định của NHNN;\n"
            "- Xác định chi phí đáp ứng nhu cầu thanh khoản trong định giá vốn nội bộ;\n"
            "- Đa dạng hóa nguồn vốn huy động về cơ cấu, kỳ hạn."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "48",
        "nghia_vu": (
            "Điều 48. Kiểm tra sức chịu đựng về thanh khoản\n"
            "Ngân hàng phải áp dụng tối thiểu 03 kịch bản:\n"
            "a) Kịch bản đặc thù ngân hàng (khủng hoảng thanh khoản riêng ngân hàng);\n"
            "b) Kịch bản toàn hệ thống (khủng hoảng tài chính quy mô rộng);\n"
            "c) Kịch bản hỗn hợp (kết hợp cả hai).\n"
            "Căn cứ kết quả, phải lập kế hoạch dự phòng thanh khoản: nhận diện khủng hoảng, thẩm quyền kích hoạt, xác định nguồn dự phòng, cơ chế báo cáo; kế hoạch rà soát định kỳ."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Thực hiện stress testing thanh khoản với tối thiểu 03 kịch bản theo quy định;\n"
            "- Xây dựng kế hoạch dự phòng thanh khoản đầy đủ theo Khoản 4 Điều 48;\n"
            "- Kế hoạch dự phòng phải xác định rõ thẩm quyền kích hoạt và nguồn dự phòng;\n"
            "- Rà soát và cập nhật kế hoạch dự phòng thanh khoản định kỳ."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "53",
        "nghia_vu": (
            "Điều 53. Chiến lược và hạn mức rủi ro lãi suất trên sổ ngân hàng (IRRBB)\n"
            "Ngân hàng phải ban hành chiến lược QLRR IRRBB, tối thiểu:\n"
            "a) Nguyên tắc quản lý IRRBB theo 2 chỉ số: thay đổi thu nhập lãi thuần (ΔNII) và thay đổi giá trị kinh tế vốn chủ sở hữu (ΔEVE);\n"
            "b) Nguyên tắc phòng ngừa IRRBB.\n"
            "Hạn mức IRRBB tối thiểu: hạn mức ΔNII và ΔEVE.\n"
            "(Từ 01/07/2026-31/12/2027: được sử dụng 1 trong 2 chỉ tiêu; từ 01/01/2028: phải dùng cả 2)"
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành chiến lược QLRR IRRBB với nguyên tắc quản lý theo ΔNII và ΔEVE;\n"
            "- Thiết lập hạn mức ΔNII và ΔEVE theo quy định;\n"
            "- Từ 01/07/2026: có thể sử dụng 1 trong 2 chỉ tiêu;\n"
            "- Từ 01/01/2028: phải đo lường và kiểm tra sức chịu đựng theo cả ΔNII lẫn ΔEVE;\n"
            "- Tính toán ΔNII và ΔEVE theo Phụ lục V Thông tư."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "56",
        "nghia_vu": (
            "Điều 56. Yêu cầu quản lý rủi ro mô hình (từ 01/01/2028)\n"
            "Ngân hàng phải:\n"
            "a) Xác định danh mục mô hình; xếp hạng mô hình (cao/trung bình/thấp);\n"
            "b) Quản lý rủi ro mô hình theo vòng đời (xây dựng, triển khai, sử dụng, theo dõi, kiểm định);\n"
            "c) Mô hình xếp hạng nội bộ phải được xác định là mô hình có rủi ro cao;\n"
            "d) Kiểm định mô hình ít nhất hàng năm và khi có thay đổi trọng yếu;\n"
            "e) Lập, quản lý hồ sơ mô hình cho tối thiểu các mô hình có rủi ro cao;\n"
            "f) Có 03 tuyến bảo vệ độc lập cho quản lý rủi ro mô hình."
        ),
        "hanh_dong": (
            "TCB phải (trước 01/01/2028):\n"
            "- Xây dựng danh mục mô hình và phân loại theo mức độ rủi ro (cao/trung bình/thấp);\n"
            "- Mô hình xếp hạng nội bộ phải được quản lý như mô hình có rủi ro cao;\n"
            "- Thiết lập quy trình quản lý theo vòng đời mô hình;\n"
            "- Thực hiện kiểm định mô hình định kỳ hàng năm và đột xuất khi có thay đổi trọng yếu;\n"
            "- Lưu hồ sơ mô hình cho tất cả mô hình có rủi ro cao (bao gồm mô hình đang dùng, đã điều chỉnh, đã ngừng);\n"
            "- Thiết lập 03 tuyến bảo vệ độc lập cho quản lý rủi ro mô hình."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "59",
        "nghia_vu": (
            "Điều 59-63. Đánh giá nội bộ về mức đủ vốn (ICAAP)\n"
            "Ngân hàng phải thực hiện ICAAP hàng năm và đột xuất. Nội dung ICAAP gồm:\n"
            "a) Đo lường rủi ro, xác định vốn kinh tế theo kế hoạch kinh doanh;\n"
            "b) Kiểm tra sức chịu đựng về vốn (kịch bản bình thường và bất lợi);\n"
            "c) Xác định vốn mục tiêu và vốn tự có dự kiến;\n"
            "d) Lập kế hoạch vốn (phương án tăng vốn, chính sách chia cổ tức, phân bổ vốn theo loại rủi ro);\n"
            "e) Giám sát mức đủ vốn;\n"
            "f) Rà soát quy trình ICAAP tối thiểu hàng năm bởi bộ phận độc lập;\n"
            "g) Kế hoạch vốn do HĐQT phê duyệt."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Thực hiện ICAAP tối thiểu hàng năm cho giai đoạn 03-05 năm tới;\n"
            "- Kiểm tra sức chịu đựng về vốn theo 02 kịch bản (bình thường và bất lợi) tối thiểu hàng năm;\n"
            "- HĐQT phê duyệt kế hoạch vốn theo đề nghị của TGĐ;\n"
            "- Kế hoạch vốn phải có phương án tăng vốn khả thi và chính sách chia cổ tức phù hợp;\n"
            "- Rà soát quy trình ICAAP hàng năm bởi bộ phận độc lập;\n"
            "- Lập báo cáo nội bộ về ICAAP hàng năm và gửi NHNN trong 90 ngày."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "64",
        "nghia_vu": (
            "Điều 64. Nguyên tắc kiểm toán nội bộ\n"
            "Kiểm toán nội bộ phải tuân thủ 03 nguyên tắc:\n"
            "1. Độc lập: Không đảm nhận công việc của tuyến 1 và 2; không chịu chi phối của tuyến 1 và 2; KTVNB không kiểm toán hoạt động mình từng phụ trách trong 03 năm; lương/thưởng tách biệt với kết quả kinh doanh của tuyến 1 và 2;\n"
            "2. Khách quan: Ghi nhận kiểm toán phải dựa trên dữ liệu, thông tin; trung thực khi đánh giá;\n"
            "3. Chuyên nghiệp: Bộ phận KTNB phải có ít nhất 01 kiểm toán viên công nghệ (có bằng CNTT, ≥2 năm kinh nghiệm CNTT)."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Đảm bảo bộ phận KTNB hoàn toàn độc lập với tuyến 1 và tuyến 2;\n"
            "- Tiêu chí lương, thưởng của KTVNB phải tách biệt với kết quả kinh doanh của đơn vị được kiểm toán;\n"
            "- KTVNB không kiểm toán hoạt động mình từng phụ trách trong vòng 03 năm;\n"
            "- Bộ phận KTNB phải có ít nhất 01 KTVNB công nghệ (CNTT, ≥2 năm kinh nghiệm);\n"
            "- Trưởng KTNB báo cáo BKS kịp thời khi phát hiện vi phạm các nguyên tắc kiểm toán."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "68",
        "nghia_vu": (
            "Điều 68. Quy định nội bộ về kiểm toán nội bộ\n"
            "Quy định nội bộ của Ban kiểm soát phải có quy định về KTNB tối thiểu:\n"
            "1. Cơ cấu tổ chức, nhiệm vụ, quyền hạn của bộ phận KTNB; tiêu chuẩn KTVNB; chuẩn mực đạo đức;\n"
            "2. Tiêu chí xác định mức độ rủi ro, trọng yếu và tần suất kiểm toán;\n"
            "3. Quy trình lập, thực hiện kế hoạch KTNB;\n"
            "4. Rà soát, đánh giá quy định KTNB; xử lý kiến nghị của NHNN;\n"
            "5. Quy định về thuê chuyên gia, tổ chức bên ngoài;\n"
            "6. Chế độ báo cáo nội bộ về KTNB."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- BKS ban hành quy định nội bộ về KTNB đầy đủ 06 nội dung tối thiểu;\n"
            "- Quy định tiêu chuẩn KTVNB theo Điều 66 (bằng đại học trở lên, ≥2 năm kinh nghiệm; Trưởng KTNB ≥3 năm);\n"
            "- BKS ban hành chuẩn mực đạo đức nghề nghiệp cho thành viên BKS và KTVNB;\n"
            "- Quy định rõ tiêu chí đánh giá mức độ rủi ro, tần suất kiểm toán từng bộ phận/hoạt động."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "69",
        "nghia_vu": (
            "Điều 69. Kế hoạch kiểm toán nội bộ\n"
            "Kế hoạch KTNB hằng năm phải:\n"
            "a) Được BKS ban hành trước ngày 15/12 của năm trước;\n"
            "b) Định hướng theo rủi ro: hoạt động rủi ro cao được ưu tiên, kiểm toán ít nhất mỗi năm một lần;\n"
            "c) Toàn diện: tất cả hoạt động, quy trình đều phải được kiểm toán; hoạt động trọng yếu ít nhất mỗi năm một lần;\n"
            "d) Có dự phòng nguồn lực cho KTNB đột xuất;\n"
            "e) Gửi kế hoạch KTNB cho NHNN trong 10 ngày làm việc kể từ ngày ban hành."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- BKS ban hành kế hoạch KTNB trước 15/12 hàng năm;\n"
            "- Kế hoạch phải ưu tiên kiểm toán các bộ phận/hoạt động rủi ro cao ít nhất mỗi năm;\n"
            "- Gửi kế hoạch KTNB cho NHNN trong 10 ngày làm việc sau khi ban hành;\n"
            "- Điều chỉnh kế hoạch khi có thay đổi trọng yếu về quy mô hoạt động, trạng thái rủi ro."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "74",
        "nghia_vu": (
            "Điều 74. Hiệu lực thi hành\n"
            "1. Thông tư có hiệu lực từ 01/07/2026, bãi bỏ Thông tư 13/2018/TT-NHNN và Thông tư 40/2018/TT-NHNN;\n"
            "2. Từ 01/01/2028: kiểm tra sức chịu đựng rủi ro tín dụng, thị trường (Điều 23) và quản lý dữ liệu rủi ro (Điều 24);\n"
            "3. Quản lý rủi ro mô hình: ngân hàng đã chuyển đổi IRB trước 01/07/2026 phải áp dụng từ 01/07/2026; ngân hàng khác từ 01/01/2028;\n"
            "4. Hạn mức, đo lường IRRBB: từ 01/07/2026-31/12/2027 có thể dùng 1 trong 2 chỉ tiêu; từ 01/01/2028 phải dùng cả ΔNII và ΔEVE;\n"
            "5. Nếu thực hiện sớm các điều khoản trên, phải thông báo cho NHNN trong 10 ngày."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Lập lộ trình tuân thủ Thông tư 83/2025 với mốc quan trọng: 01/07/2026 (hiệu lực chung), 01/01/2028 (stress testing tín dụng/thị trường, quản lý dữ liệu rủi ro, quản lý rủi ro mô hình toàn diện, IRRBB theo cả 2 chỉ tiêu);\n"
            "- Rà soát lộ trình chuyển đổi IRB để xác định thời điểm áp dụng quản lý rủi ro mô hình;\n"
            "- Nếu thực hiện sớm các điều khoản trễ hiệu lực, thông báo NHNN trong 10 ngày;\n"
            "- Từ 01/07/2026: dừng áp dụng Thông tư 13/2018 và Thông tư 40/2018."
        ),
        "bat_buoc": "x",
    },
]

# === Build Excel ===
# Row 1: Title (merged A1:G1)
ws.merge_cells("A1:G1")
ws["A1"] = "Nghĩa vụ pháp luật"
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Row 2: Headers
headers = [
    "Tên văn bản",
    "Số văn bản",
    "Số điều khoản",
    "Ngày hiệu lực",
    "Nghĩa vụ pháp luật",
    "Hành động Techcombank cần thực hiện",
    "Có bắt buộc hay không",
]
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF", size=11)
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row 3+: Data
row_fill_alt = PatternFill("solid", fgColor="DCE6F1")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for i, ob in enumerate(obligations):
    row_idx = i + 3
    fill = row_fill_alt if i % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")

    data = [
        ten_van_ban,
        so_van_ban,
        int(ob["dieu"]),
        ngay_hieu_luc,
        ob["nghia_vu"],
        ob["hanh_dong"],
        ob["bat_buoc"],
    ]

    for col_idx, value in enumerate(data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_idx == 4 and isinstance(value, date):
            cell.number_format = "DD/MM/YYYY"
        if col_idx == 7:
            cell.alignment = Alignment(horizontal="center", vertical="top")

# Apply border to header row
for col_idx in range(1, 8):
    ws.cell(row=2, column=col_idx).border = thin_border

# Column widths
col_widths = [30, 18, 14, 14, 55, 55, 15]
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Row heights
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 30
for i in range(len(obligations)):
    ws.row_dimensions[i + 3].height = 120

# Freeze top 2 rows
ws.freeze_panes = "A3"

output_path = "/Users/tramynguyen/Work/Law_to_excel/Danh mục nghĩa vụ tuân thủ - TT83.2025.TT.NHNN.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
