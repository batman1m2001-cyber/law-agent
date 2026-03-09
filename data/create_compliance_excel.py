import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Trang tính1"

# === Data ===
ten_van_ban = "Thông tư về hoạt động tư vấn của tổ chức tín dụng, chi nhánh ngân hàng nước ngoài"
so_van_ban = "38/2024/TT-NHNN"
ngay_hieu_luc = date(2024, 7, 1)

obligations = [
    {
        "dieu": "4",
        "nghia_vu": (
            "Điều 4. Phạm vi hoạt động tư vấn\n"
            "Tổ chức tín dụng cung ứng dịch vụ tư vấn chỉ được cung ứng dịch vụ tư vấn về "
            "hoạt động ngân hàng, hoạt động kinh doanh khác quy định trong Giấy phép do NHNN cấp. "
            "Giải pháp, phương án đề xuất phải được thực hiện bởi tổ chức tín dụng có Giấy phép phù hợp."
        ),
        "hanh_dong": (
            "TCB chỉ được cung ứng dịch vụ tư vấn trong phạm vi hoạt động được ghi nhận tại Giấy phép do NHNN cấp. "
            "Không được tư vấn vượt phạm vi Giấy phép. Giải pháp đề xuất phải do TCB hoặc tổ chức tín dụng có Giấy phép phù hợp thực hiện."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "5",
        "nghia_vu": (
            "Điều 5. Nguyên tắc hoạt động tư vấn\n"
            "1. Tuân thủ pháp luật và chịu trách nhiệm về hoạt động tư vấn của nhân viên.\n"
            "2. Giám sát việc tuân thủ đạo đức nghề nghiệp của nhân viên tư vấn.\n"
            "3. Đảm bảo tính độc lập, trung thực, khách quan.\n"
            "4. Bảo mật thông tin khách hàng (trừ trường hợp có thỏa thuận khác hoặc pháp luật quy định khác).\n"
            "5. Tự chủ trong hoạt động tư vấn.\n"
            "6. Tư vấn bằng phương tiện điện tử phải tuân thủ pháp luật về giao dịch điện tử."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Xây dựng cơ chế chịu trách nhiệm về nội dung tư vấn của nhân viên;\n"
            "- Thiết lập hệ thống giám sát tuân thủ đạo đức nghề nghiệp;\n"
            "- Đảm bảo quy trình tư vấn độc lập, khách quan;\n"
            "- Thực hiện bảo mật thông tin khách hàng theo quy định;\n"
            "- Đảm bảo hạ tầng tư vấn điện tử tuân thủ pháp luật giao dịch điện tử."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "6",
        "nghia_vu": (
            "Điều 6. Quy định nội bộ\n"
            "Trước khi triển khai hoạt động tư vấn, tổ chức tín dụng phải xây dựng quy định nội bộ về hoạt động tư vấn. "
            "Trong vòng 10 ngày kể từ ngày ban hành hoặc sửa đổi, bổ sung quy định nội bộ, phải gửi về NHNN (Cơ quan Thanh tra, giám sát ngân hàng).\n"
            "Quy định nội bộ phải có tối thiểu:\n"
            "a) Quy trình tiếp nhận yêu cầu khách hàng và ký hợp đồng;\n"
            "b) Quy trình thực hiện tư vấn và quản lý rủi ro;\n"
            "c) Trách nhiệm, nghĩa vụ từng bộ phận/cá nhân; những việc nhân viên được/không được làm;\n"
            "d) Tiêu chuẩn đạo đức nghề nghiệp và cơ chế giám sát;\n"
            "đ) Quy định đào tạo, kiểm tra chất lượng chuyên môn định kỳ;\n"
            "e) Quy trình tiếp nhận và giải quyết khiếu nại khách hàng."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Ban hành quy định nội bộ về hoạt động tư vấn trước khi triển khai dịch vụ;\n"
            "- Gửi quy định nội bộ (và mọi lần sửa đổi) đến NHNN (Cơ quan Thanh tra, giám sát) trong vòng 10 ngày;\n"
            "- Đảm bảo quy định nội bộ có đầy đủ 6 nội dung bắt buộc theo khoản 2 Điều 6."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "7",
        "nghia_vu": (
            "Điều 7. Tiêu chuẩn đạo đức nghề nghiệp của nhân viên tư vấn\n"
            "1. Có hiểu biết chuyên môn về nội dung tư vấn.\n"
            "2. Trung thực, công bằng, cẩn trọng, liêm chính; không cung cấp thông tin sai hoặc gây hiểu lầm.\n"
            "3. Vì lợi ích khách hàng; nội dung tư vấn phù hợp với thông tin do khách hàng cung cấp.\n"
            "4. Giữ an toàn, tách biệt tiền/tài sản khách hàng với giao dịch được tư vấn.\n"
            "5. Không quyết định thay khách hàng; không chuyển thông tin khách hàng cho bên thứ ba khi chưa được đồng ý.\n"
            "6. Cảnh báo khách hàng về các rủi ro có thể phát sinh."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Đảm bảo nhân viên tư vấn có đủ năng lực chuyên môn trước khi phân công;\n"
            "- Đào tạo và kiểm tra định kỳ về đạo đức nghề nghiệp;\n"
            "- Ban hành quy định cấm nhân viên cung cấp thông tin sai, thông tin gây hiểu lầm;\n"
            "- Có cơ chế tách biệt tiền/tài sản khách hàng;\n"
            "- Quy định rõ trách nhiệm bảo mật thông tin khách hàng;\n"
            "- Yêu cầu nhân viên cảnh báo rủi ro cho khách hàng trước khi tư vấn."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "8",
        "nghia_vu": (
            "Điều 8. Tiếp xúc khách hàng\n"
            "Trường hợp nhân viên tư vấn có ý định tư vấn cho khách hàng về một sản phẩm, dịch vụ, giao dịch "
            "khi chưa được khách hàng yêu cầu, nhân viên tư vấn phải thông báo và phải được khách hàng đồng ý trước khi thực hiện tư vấn."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Quy định rõ trong quy trình nội bộ: khi nhân viên chủ động tư vấn (không có yêu cầu từ khách hàng) "
            "phải thông báo và được khách hàng đồng ý bằng văn bản hoặc hình thức phù hợp trước khi tiến hành tư vấn;\n"
            "- Không thực hiện tư vấn khi chưa có sự đồng ý của khách hàng."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "9",
        "nghia_vu": (
            "Điều 9. Phí trong hoạt động tư vấn\n"
            "Tổ chức tín dụng có quyền thu phí trong hoạt động tư vấn. "
            "Mức phí phải được niêm yết công khai và phải được ghi trong hợp đồng tư vấn."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Niêm yết công khai biểu phí dịch vụ tư vấn tại quầy giao dịch và/hoặc website;\n"
            "- Ghi rõ mức phí và phương thức thanh toán trong hợp đồng tư vấn với từng khách hàng."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "10",
        "nghia_vu": (
            "Điều 10. Hợp đồng tư vấn\n"
            "Phải thống nhất thỏa thuận bằng hợp đồng tư vấn. Hợp đồng tối thiểu phải có:\n"
            "a) Thông tin các bên (tên, địa chỉ, người đại diện, số định danh cá nhân...);\n"
            "b) Mục tiêu, phạm vi tư vấn;\n"
            "c) Phương thức thực hiện tư vấn;\n"
            "d) Quyền và nghĩa vụ của tổ chức tín dụng;\n"
            "đ) Quyền và nghĩa vụ của khách hàng;\n"
            "e) Thời hạn thực hiện hợp đồng;\n"
            "g) Phí và phương thức thanh toán;\n"
            "h) Phương thức giải quyết tranh chấp;\n"
            "i) Hiệu lực hợp đồng;\n"
            "k) Thời hạn hợp đồng;\n"
            "l) Thỏa thuận bồi thường thiệt hại khi vi phạm."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Xây dựng mẫu hợp đồng tư vấn đảm bảo đầy đủ 12 nội dung bắt buộc theo khoản 2 Điều 10;\n"
            "- Ký hợp đồng tư vấn với khách hàng trước khi cung ứng dịch vụ;\n"
            "- Thực hiện sửa đổi, bổ sung hợp đồng theo đúng quy định pháp luật khi có thay đổi."
        ),
        "bat_buoc": "x",
    },
    {
        "dieu": "11",
        "nghia_vu": (
            "Điều 11. Hiệu lực thi hành\n"
            "Thông tư có hiệu lực từ ngày 01/07/2024. "
            "Tổ chức tín dụng đã được cấp Giấy phép hoạt động tư vấn ngân hàng, tài chính theo Luật cũ "
            "được tiếp tục thực hiện theo Thông tư này mà không phải đề nghị cấp lại Giấy phép."
        ),
        "hanh_dong": (
            "TCB phải:\n"
            "- Rà soát lại Giấy phép hiện tại để xác nhận phạm vi được tư vấn;\n"
            "- Đảm bảo toàn bộ hoạt động tư vấn tuân thủ Thông tư 38/2024 kể từ ngày 01/07/2024;\n"
            "- Không cần xin cấp lại Giấy phép nếu đã được cấp phép tư vấn theo quy định cũ."
        ),
        "bat_buoc": "x",
    },
]

# === Build Excel ===
# Row 1: Title (merged A1:G1)
ws.merge_cells("A1:G1")
ws["A1"] = "Nghĩa vụ pháp luật"
ws["A1"].font = Font(bold=True, size=13)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Row 2: Headers
headers = [
    "Tên văn bản",
    "Số văn bản",
    "Số điều khoản",
    "Ngày hiệu lực",
    "Nghĩa vụ pháp luật",
    "Hành động Techcombank cần thực hiện",
    "Có bắt buộc hay không",
]
header_fill = PatternFill("solid", fgColor="4472C4")
header_font = Font(bold=True, color="FFFFFF", size=11)
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Row 3+: Data
row_fill_alt = PatternFill("solid", fgColor="DCE6F1")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for i, ob in enumerate(obligations):
    row_idx = i + 3
    fill = row_fill_alt if i % 2 == 1 else PatternFill("solid", fgColor="FFFFFF")

    data = [
        ten_van_ban,
        so_van_ban,
        int(ob["dieu"]),
        ngay_hieu_luc,
        ob["nghia_vu"],
        ob["hanh_dong"],
        ob["bat_buoc"],
    ]

    for col_idx, value in enumerate(data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if col_idx == 4 and isinstance(value, date):
            cell.number_format = "DD/MM/YYYY"
        if col_idx == 7:
            cell.alignment = Alignment(horizontal="center", vertical="top")

# Apply border to header row
for col_idx in range(1, 8):
    ws.cell(row=2, column=col_idx).border = thin_border

# Column widths
col_widths = [30, 18, 14, 14, 55, 55, 15]
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Row heights
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 30
for i in range(len(obligations)):
    ws.row_dimensions[i + 3].height = 120

# Freeze top 2 rows
ws.freeze_panes = "A3"

output_path = "/Users/tramynguyen/Work/Law_to_excel/Danh mục nghĩa vụ tuân thủ - TT38.2024.TT.NHNN.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
