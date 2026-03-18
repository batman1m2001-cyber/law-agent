"""Export obligations from JSON to formatted Excel.

Usage:
  uv run python -m law_agent.export <output.json> [result.xlsx]
"""

import json
import re
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "Tên văn bản", "Số văn bản", "Số điều khoản", "Ngày hiệu lực",
    "Nghĩa vụ pháp luật", "Hành động Techcombank cần thực hiện",
    "Có bắt buộc hay không", "Phân loại", "Chủ thể tương tác", "Chủ thể hoạt động",
]
LOAI_LABEL = {"bat_buoc": "Nghĩa vụ", "quyen": "Quyền", "dinh_nghia": "Quy định chung"}
COL_WIDTHS = [30, 18, 14, 14, 55, 55, 15, 18, 25, 25]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
KHOAN_RE = re.compile(r"^(\d+)\.\s", re.MULTILINE)


def _extract_khoan_text(article_text: str, dieu: str) -> str:
    parts = dieu.split(".")
    if len(parts) < 2:
        return article_text
    khoan_num = parts[1]
    matches = list(KHOAN_RE.finditer(article_text))
    target = None
    for m in matches:
        if m.group(1) == khoan_num:
            target = m
            break
    if target is None:
        return article_text
    start = target.start()
    end = len(article_text)
    for m in matches:
        if m.start() > start:
            end = m.start()
            break
    return article_text[start:end].strip()


def _parse_bold(text: str) -> CellRichText | str:
    if "**" not in text:
        return text
    bold_font = InlineFont(b=True, sz=11)
    normal_font = InlineFont(sz=11)
    parts = []
    for i, seg in enumerate(re.split(r"\*\*", text)):
        if seg:
            parts.append(TextBlock(bold_font if i % 2 == 1 else normal_font, seg))
    return CellRichText(*parts) if parts else text


def export(json_path: str, excel_path: str | None = None):
    """Read JSON, export all 'done' articles to Excel."""
    data = json.loads(Path(json_path).read_text(encoding="utf-8"))
    meta = data.get("metadata", {})
    ten_van_ban = meta.get("ten_van_ban", "")
    so_van_ban = meta.get("so_van_ban", "")

    ngay_hieu_luc = None
    nhl = meta.get("ngay_hieu_luc", "")
    if nhl:
        p = nhl.split("-")
        if len(p) == 3:
            try:
                ngay_hieu_luc = date(int(p[0]), int(p[1]), int(p[2]))
            except ValueError:
                pass

    rows = []
    for art_key, art in data.get("articles", {}).items():
        if art.get("status") != "done":
            continue
        title = art.get("title", "")
        text = art.get("text", "")
        for ob in art.get("obligations", []):
            dieu = ob.get("dieu", "")
            khoan_text = _extract_khoan_text(text, dieu)
            rows.append({
                "ten_van_ban": ten_van_ban,
                "so_van_ban": so_van_ban,
                "dieu": dieu,
                "ngay_hieu_luc": ngay_hieu_luc,
                "nghia_vu": f"Điều {dieu}. {title}\n{khoan_text}" if title else khoan_text,
                "hanh_dong": ob.get("hanh_dong", ""),
                "bat_buoc": "x" if ob.get("loai") == "bat_buoc" else "",
                "phan_loai": LOAI_LABEL.get(ob.get("loai", ""), "Không"),
                "chu_the_tuong_tac": ob.get("chu_the_tuong_tac", ""),
                "chu_the_hoat_dong": ob.get("chu_the_hoat_dong", ""),
            })

    if not rows:
        print("No obligations found.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Nghĩa vụ tuân thủ"

    ws.merge_cells("A1:J1")
    ws["A1"].value = "Nghĩa vụ pháp luật"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for i, row in enumerate(rows):
        r = i + 3
        fill = ALT_FILL if i % 2 == 1 else WHITE_FILL
        vals = [
            row["ten_van_ban"], row["so_van_ban"], row["dieu"], row["ngay_hieu_luc"],
            row["nghia_vu"], row["hanh_dong"], row["bat_buoc"], row["phan_loai"],
            row["chu_the_tuong_tac"], row["chu_the_hoat_dong"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if c == 4 and isinstance(v, date):
                cell.number_format = "DD/MM/YYYY"
            if c == 6 and isinstance(v, str) and v:
                cell.value = _parse_bold(v)
            if c in (7, 8):
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    for c, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    for i in range(len(rows)):
        ws.row_dimensions[i + 3].height = 120
    ws.freeze_panes = "A3"

    if not excel_path:
        excel_path = json_path.rsplit(".", 1)[0] + ".xlsx"
    wb.save(excel_path)
    print(f"Excel exported: {excel_path} ({len(rows)} obligations)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: uv run python -m law_agent.export <output.json> [result.xlsx]")
        sys.exit(1)
    export(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
