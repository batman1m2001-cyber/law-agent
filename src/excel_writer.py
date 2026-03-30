"""Create formatted Excel output from extracted obligations."""

import io
import re
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .extractor import Obligation

HEADERS = [
    "Tên văn bản",
    "Số văn bản",
    "Số điều khoản",
    "Ngày hiệu lực",
    "Nghĩa vụ pháp luật",
    "Hành động Techcombank cần thực hiện",
    "Có bắt buộc hay không",
    "Chủ thể tương tác",
    "Chủ thể hoạt động",
    "Nhóm nghĩa vụ",
    "Lý do nhóm nghĩa vụ",
]

COL_WIDTHS = [30, 18, 14, 14, 55, 55, 15, 25, 25, 25, 45]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
def _resolve_chu_the_hoat_dong(ob) -> str:
    """Default chu_the_hoat_dong to 'TCB' for dinh_nghia obligations."""
    if ob.loai == "dinh_nghia":
        return "TCB"
    return ob.chu_the_hoat_dong


_KEYWORD_TO_LABEL = {
    "Cơ quan nhà nước": [
        "Ngân hàng Nhà nước", "NHNN", "tổ chức kiểm toán độc lập",
        "cơ quan thanh tra", "kiểm toán nhà nước", "cơ quan có thẩm quyền",
        "cơ quan giám sát",
    ],
    "Khách hàng": [
        "khách hàng", "người gửi tiền", "người vay", "người sử dụng dịch vụ",
    ],
}


def _fix_chu_the_tuong_tac(khoan_text: str, model_value: str) -> str:
    """Bổ sung nhãn bị bỏ sót dựa trên từ khóa xuất hiện trong văn bản khoản."""
    result = model_value
    for label, keywords in _KEYWORD_TO_LABEL.items():
        if label not in result and any(kw.lower() in khoan_text.lower() for kw in keywords):
            result = (result + ", " + label).lstrip(", ")
    return result


def _dedupe_chu_the_tuong_tac(value: str) -> str:
    """If both 'Quản trị nội bộ' and 'Cơ quan nhà nước' are present, drop 'Quản trị nội bộ'."""
    if not value:
        return value
    parts = [p.strip() for p in value.split(",")]
    if "Quản trị nội bộ" in parts and "Cơ quan nhà nước" in parts:
        parts = [p for p in parts if p != "Quản trị nội bộ"]
    return ", ".join(parts)


def _format_bullet_text(text: str) -> str:
    """Insert newlines before bullet markers • and - so Excel wraps correctly."""
    if not text:
        return text
    # Newline before • (top-level bullets)
    text = re.sub(r"\s*•\s*", "\n• ", text)
    # Newline before - (sub-bullets), but only when preceded by space/text (not at start)
    text = re.sub(r"(?<=[^\n])\s+-\s+", "\n  - ", text)
    return text.strip()


def _strip_html(html: str) -> str:
    """Convert HTML to plain text for Excel cells."""
    text = re.sub(r"<br\s*/?>", "\n", html, flags=re.IGNORECASE)
    text = re.sub(r"</p>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<p[^>]*>", "", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def create_excel(
    obligations: list[Obligation],
    ten_van_ban: str,
    so_van_ban: str,
    ngay_hieu_luc: date | datetime,
) -> io.BytesIO:
    """Create a formatted Excel workbook from obligations.

    Returns a BytesIO buffer containing the .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Nghĩa vụ tuân thủ"

    # Row 1: Title (merged)
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "Nghĩa vụ pháp luật"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Headers
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Row 3+: Data
    for i, ob in enumerate(obligations):
        row_idx = i + 3
        fill = ALT_FILL if i % 2 == 1 else WHITE_FILL

        # Build nghia vu text: "Điều X. Title\nNội dung"
        noi_dung_plain = _strip_html(ob.noi_dung)
        nghia_vu_text = f"Điều {ob.dieu}. {ob.tieu_de}\n{noi_dung_plain}" if ob.tieu_de else noi_dung_plain

        bat_buoc = "x" if ob.loai == "bat_buoc" else ""

        row_data = [
            ten_van_ban,
            so_van_ban,
            ob.dieu,
            ngay_hieu_luc,
            nghia_vu_text,
            _format_bullet_text(ob.hanh_dong),
            bat_buoc,
            _dedupe_chu_the_tuong_tac(_fix_chu_the_tuong_tac(noi_dung_plain, ob.chu_the_tuong_tac)),
            _resolve_chu_the_hoat_dong(ob),
            ob.nhom_nghia_vu,
            ob.nhom_ly_do,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if col_idx == 4 and isinstance(value, (date, datetime)):
                cell.number_format = "DD/MM/YYYY"
            if col_idx == 7:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    for i in range(len(obligations)):
        ws.row_dimensions[i + 3].height = 120

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
